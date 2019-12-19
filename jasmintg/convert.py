import os
import re
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl import Workbook
import textgrids


def transform_textgrids(oArgs, errHandle):
    """Transform the textgrids in the input directory"""

    # Initialisations
    src_ext = ".TextGrid"
    outfile_name = "jasmintg.xlsx"
    headers = ['child', 'tier1', 'tier2', 'tier5', 'tier6_L', 'tier6_N']

    try:
        dirInput = oArgs['input']
        dirOutput = oArgs['output']
        force = oArgs['force']
        debug = oArgs['debug']

        # Determine the output file
        outfile = os.path.join(dirOutput, outfile_name)

        # Start a workbook
        wb = openpyxl.Workbook()
        ws = wb.get_active_sheet()
        ws.title = "Data"

        # Set up the column headers
        for col_num in range(len(headers)):
            c = ws.cell(row=1, column=col_num+1)
            c.value = headers[col_num]
            c.font = openpyxl.styles.Font(bold=True)
            # Set width to a fixed size
            ws.column_dimensions[get_column_letter(col_num+1)].width = 8.0        
            
        # Walk all the files in the input
        lst_src = [os.path.join(dirInput, f) for f in os.listdir(dirInput) if os.path.isfile(os.path.join(dirInput, f)) and src_ext in f]

        row_num = 1
        for file in lst_src:
            row_num += 1
            # Get the name of the child from the name of the file
            child = os.path.basename(file).replace(src_ext, "")


            # Show where we are
            errHandle.Status("child = {}".format(child))

            # Read the textgrid file
            grid = textgrids.TextGrid(file)

            # Access the tiers that we need
            counter = 1
            for k,v in grid.items():
                if counter == 1:
                    tier1 = v
                elif counter == 2:
                    tier2 = v
                elif counter == 5:
                    tier5 = v
                elif counter == 6:
                    tier6 = v

                counter += 1

            offset2 = 0
            offset5 = 0
            offset6 = 0

            # Walk through all the items in tier1
            for idx, t1 in enumerate(tier1):

                # Get the corresponding values in the other tiers
                t2 = tier2[idx+offset2]
                t5 = tier5[idx+offset5]
                t6 = tier6[idx+offset6]

                while not is_close(t1,t2) and t2.xmin < t1.xmin and idx+offset2  < len(tier2) - 1:
                    offset2 += 1
                    t2 = tier2[idx+offset2]
                while not is_close(t1,t2) and t2.xmin > t1.xmin:
                    offset2 -= 1
                    t2 = tier2[idx+offset2]

                while not is_close(t1,t5) and t5.xmin < t1.xmin and  idx+offset5  < len(tier5) - 1:
                    offset5 += 1
                    t5 = tier5[idx+offset5]
                while not is_close(t1,t5) and t5.xmin > t1.xmin:
                    offset5 -= 1
                    t5 = tier5[idx+offset5]

                while not is_close(t1,t6) and t6.xmin < t1.xmin and  idx+offset6  < len(tier6) - 1:
                    offset6 += 1
                    t6 = tier6[idx+offset6]
                while not is_close(t1,t6) and t6.xmin > t1.xmin:
                    offset6 -= 1
                    t6 = tier6[idx+offset6]


                # Check if all tiers synchronize
                if is_close(t1,t2) and is_close(t1, t5) and is_close(t1, t6):
                    # All is well: process
                    arCombi = t6.text.split("/")
                    letter = ""
                    number = ""
                    if len(arCombi) == 1:
                        v = arCombi[0]
                        if re.match("^\d+$", v):
                            number = v
                            letter = ""
                        else:
                            number = ""
                            letter = v
                    elif len(arCombi) == 2:
                        letter = arCombi[0].strip()
                        number = arCombi[1].strip()

                    # Create list of values
                    row = [child, t1.text, t2.text, t5.text, letter, number]

                    for idx, v in enumerate(row):
                        cell_this = ws.cell(row=row_num, column=idx+1)
                        cell_this.value = v
                        cell_this.alignment = openpyxl.styles.Alignment(wrap_text=False)

                    # We are going to the next row
                    row_num += 1

                else:
                    # Synchronization problem
                    msg = "Synchronization problem in [{}] tier {} t1={} t2={} t5={} t6={}".format(child, idx, t1.xmin, t2.xmin, t5.xmin, t6.xmin)
                    errHandle.Status(msg)
            
        # Save the result
        wb.save(outfile)

        return True
    except:
        errHandle.DoError("transform_textgrids")
        return False


def is_close(t1, t2):
    t1xmin = t1.xmin
    t2xmin = t2.xmin
    diff = abs(t1xmin - t2xmin)
    bFound = (diff < 0.3)
    return bFound